import time
import cv2
import numpy as np
from PyQt5.QtCore import QTimer, Qt, QSize, pyqtSignal, QObject
from PyQt5.QtGui import QImage, QPixmap, QMovie, QIntValidator
from PyQt5.QtWidgets import (QVBoxLayout, QHBoxLayout, QLabel, QSpacerItem,
                             QPushButton, QScrollArea, QWidget, QSizePolicy,
                             QGroupBox, QMessageBox, QApplication, QDialog, QTabWidget, QLineEdit, QComboBox,
                             QDialogButtonBox)
from classes.report_dialog import ReportDialog
from classes.thresholds_editor import ThresholdsEditor
from classes.squares_widget import SquaresWidget
from utils.excel_processing import generate_report
from pathlib import Path
from classes.live_camera_manager import LiveCameraManager
from classes.squares_widget import STATE_COLORS
from utils.auto_move import send_command
import time
from PyQt5.QtWidgets import QApplication
import matplotlib
from processing.draw_contour import draw_part_contour
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import sys
from PyQt5.QtWidgets import QMainWindow
from PyQt5.QtCore import Qt

class ThresholdsDialog(QDialog):
    def __init__(self, cam_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Пороговые значения брака")
        self.setModal(True)
        layout = QVBoxLayout()
        self.editor = ThresholdsEditor(cam_manager)
        layout.addWidget(self.editor.group)
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        self.setLayout(layout)


class BatchNameDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Новая партия")
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Введите номер/название партии:"))
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Например: Партия_001")
        layout.addWidget(self.name_edit)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.setLayout(layout)

    def get_name(self):
        return self.name_edit.text().strip()

class StatsDialog(QDialog):
    def __init__(self, cam_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Статистика по изделиям")
        self.setModal(True)
        self.cam_manager = cam_manager

        self.stats_group = self.create_stats_widget()

        layout = QVBoxLayout()
        layout.addWidget(self.stats_group)
        btn_close = QPushButton("Закрыть")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        self.setLayout(layout)

    def create_stats_widget(self):
        stats_group = QGroupBox("Статистика по изделиям")
        stats_group.setFixedSize(350, 284)
        stats_group.setStyleSheet("""
            QGroupBox {
                font-family: "Montserrat";
                font-weight: bold;
                font-size: 20px;
                color: #343e43;
                background-color: #f5f5f5;
                border: 4px solid #d9d9d9;
                border-radius: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                top: -4px;
                padding: 0 16px;
                background-color: #f5f5f5;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(12, 32, 12, 12)

        stats = self.cam_manager.stats_manager.get_stats_snapshot()

        good_label = QLabel("Годные изделия")
        bad_label = QLabel("Бракованные изделия")
        checked_label = QLabel("Всего проверено")

        for label in [good_label, bad_label, checked_label]:
            label.setFixedSize(160, 36)
            label.setStyleSheet("""
                QLabel {
                    font-family: "Montserrat";
                    font-size: 14px;
                    font-weight: bold;
                    color: #343e43;
                }
            """)

        good_amount = QLabel(str(stats["total_good"]))
        bad_amount = QLabel(str(stats["total_bad"]))
        checked_amount = QLabel(str(stats["total_checked"]))

        for amount in [good_amount, bad_amount, checked_amount]:
            amount.setFixedSize(70, 36)
            amount.setAlignment(Qt.AlignmentFlag.AlignCenter)
            amount.setStyleSheet("""
                QLabel {
                    background-color: #d9d9d9;
                    border: 2px solid #343e43;
                    border-radius: 12px;
                    font-family: "Montserrat";
                    font-size: 16px;
                    font-weight: bold;
                    color: #343e43;
                    padding: 8px 12px;
                }
            """)

        def add_row(row_layout, label, amount):
            row_layout.addWidget(label)
            row_layout.addStretch()
            row_layout.addWidget(amount)

        good_layout = QHBoxLayout()
        add_row(good_layout, good_label, good_amount)

        bad_layout = QHBoxLayout()
        add_row(bad_layout, bad_label, bad_amount)

        checked_layout = QHBoxLayout()
        add_row(checked_layout, checked_label, checked_amount)

        clear_button = QPushButton("Очистить статистику")
        clear_button.setFixedSize(320, 48)
        clear_button.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 20px;
                border: none;
                border-radius: 12px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #8C1212;
                color: #FFFFFF;
            }
        """)
        clear_button.clicked.connect(self.reset_stats)

        layout.addLayout(good_layout)
        layout.addLayout(bad_layout)
        layout.addLayout(checked_layout)
        layout.addSpacing(10)
        layout.addWidget(clear_button)

        stats_group.setLayout(layout)
        return stats_group

    def reset_stats(self):
        self.cam_manager.stats_manager.reset_stats()
        QMessageBox.information(self, "Сброс", "Статистика сброшена.")
        self.accept()


class StatsUpdater(QObject):
    stats_updated = pyqtSignal(dict)

    def __init__(self, cam_manager):
        super().__init__()
        self.cam_manager = cam_manager

    def update_stats(self):
        try:
            stats = self.cam_manager.stats_manager.get_stats_snapshot()
            self.stats_updated.emit(stats)
        except Exception as e:
            print(f"Ошибка обновления статистики: {e}")


class CameraViewer(QWidget):
    def __init__(self, cam_manager):
        super().__init__()
        self.anim = QMovie("assets/load.gif")
        self.anim.setScaledSize(QSize(768, 768))
        self.anim.start()
        self.cam_manager = cam_manager
        self.current_cam_index = 0
        self.setWindowTitle("Пульт управления конвейером")
        self.setGeometry(100, 100, 1920, 960)
        self.setStyleSheet("background-color: #f5f5f5;")

        self.image_label = QLabel(self)
        self.image_label.setFixedSize(1280, 720)
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setStyleSheet("""
            QLabel {
                background-color: #E0E0E0;
                border-radius: 16px;
            }
        """)

        self.info_table = QLabel(self)
        self.info_table.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self.info_table.setStyleSheet("""
            QLabel {
                background-color: #f5f5f5;
                padding: 12px;
                font-family: "Montserrat";
                font-weight: semi-bold;
                font-size: 16px;
                color: #343e43;
            }
        """)

        self.scroll_group_left = QGroupBox("Журнал проверки")
        self.scroll_group_left.setFixedSize(550, 500)
        self.scroll_group_left.setStyleSheet("""
            QGroupBox {
                font-family: "Montserrat";
                font-weight: bold;
                font-size: 20px;
                color: #343e43;
                background-color: #f5f5f5;
                border: 4px solid #d9d9d9;
                border-radius: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                top: -6px;
                padding: 0 16px;
                background-color: #f5f5f5;
            }
        """)

        self.scroll_area_left = QScrollArea(self)
        self.scroll_area_left.setWidgetResizable(True)
        self.scroll_area_left.setWidget(self.info_table)
        self.scroll_area_left.setStyleSheet("QScrollArea { border: transparent; }")

        self.thresholds_editor = ThresholdsEditor(self.cam_manager)

        self.btn_prev = QPushButton("← Предыдущая камера", self)
        self.btn_next = QPushButton("Следующая камера →", self)

        for btn in [self.btn_prev, self.btn_next]:
            btn.setFixedSize(340, 56)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #D9D9D9;
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 20px;
                    border: 2px solid #000000;
                    border-radius: 12px;
                    padding: 8px;
                }
                QPushButton:disabled {
                    background-color: #B0B0B0;
                }
                QPushButton:hover {
                    background-color: #3997E4;
                    color: #FFFFFF;
                }
            """)

        self.btn_prev.clicked.connect(self.prev_camera)
        self.btn_next.clicked.connect(self.next_camera)

        self.btn_start = QPushButton("Старт", self)
        self.btn_stop = QPushButton("Стоп", self)
        self.btn_start.clicked.connect(self.start_conveyor)
        self.btn_stop.clicked.connect(self.stop_conveyor)

        self.btn_start.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 16px;
                padding: 8px;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)

        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 16px;
                padding: 8px;
            }
            QPushButton:disabled {
                background-color: #9C6363;
                color: #FFFFFF;
            }
            QPushButton:hover {
                background-color: #8C1212;
                color: #FFFFFF;
            }
        """)

        self.btn_new_batch = QPushButton("Начать новую партию", self)
        self.btn_new_batch.clicked.connect(self.start_new_batch)
        self.btn_new_batch.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 16px;
                padding: 8px;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)

        self.btn_start.setFixedSize(192, 56)
        self.btn_stop.setFixedSize(192, 56)
        self.btn_new_batch.setFixedSize(340, 56)
        self.btn_stop.setEnabled(False)

        self.status_indicator = QLabel()
        self.status_indicator.setFixedSize(32, 32)
        self.status_indicator.setStyleSheet("background-color: #8C1212; border-radius: 16px;")

        self.status_text = QLabel("Система остановлена")
        self.status_text.setStyleSheet("""
            color: #343E43;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 24px;
        """)

        self.move_text = QLabel("Перемещение ленты только при остановке инспекции")
        self.move_text.setStyleSheet("""
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 20px;
                """)

        self.btn_move_forw = QPushButton("+")
        self.btn_move_forw.setFixedSize(192, 56)
        self.btn_move_forw.setStyleSheet("""
                                    QPushButton {
                                        background-color: #D9D9D9;
                                        color: #343E43;
                                        font-family: Montserrat;
                                        font-weight: bold;
                                        font-size: 16px;
                                        border: 2px solid #000000;
                                        border-radius: 16px;
                                        padding: 8px;
                                    }
                                    QPushButton:hover {
                                        background-color: #3997E4;
                                        color: #FFFFFF;
                                    }
                                """)
        self.btn_move_forw.pressed.connect(self.on_pressed_forw)
        self.btn_move_forw.released.connect(self.on_released_forw)

        self.btn_move_back = QPushButton("-")
        self.btn_move_back.setFixedSize(192, 56)
        self.btn_move_back.setStyleSheet("""
                                    QPushButton {
                                        background-color: #D9D9D9;
                                        color: #343E43;
                                        font-family: Montserrat;
                                        font-weight: bold;
                                        font-size: 16px;
                                        border: 2px solid #000000;
                                        border-radius: 16px;
                                        padding: 8px;
                                    }
                                    QPushButton:hover {
                                        background-color: #3997E4;
                                        color: #FFFFFF;
                                    }
                                """)
        self.btn_move_back.pressed.connect(self.on_pressed_back)
        self.btn_move_back.released.connect(self.on_released_back)

        status_layout = QHBoxLayout()
        status_spacer_outer = QSpacerItem(32, 32, QSizePolicy.Expanding, QSizePolicy.Minimum)
        status_spacer_inner = QSpacerItem(12, 32, QSizePolicy.Fixed, QSizePolicy.Minimum)
        status_layout.addSpacerItem(status_spacer_outer)
        status_layout.addWidget(self.status_indicator)
        status_layout.addSpacerItem(status_spacer_inner)
        status_layout.addWidget(self.status_text)
        status_layout.addSpacerItem(status_spacer_outer)

        control_layout = QHBoxLayout()
        control_spacer_outer = QSpacerItem(16, 32, QSizePolicy.Expanding, QSizePolicy.Minimum)
        control_spacer_inner = QSpacerItem(18, 32, QSizePolicy.Fixed, QSizePolicy.Minimum)
        control_layout.addSpacerItem(control_spacer_outer)
        control_layout.addWidget(self.btn_start)
        control_layout.addSpacerItem(control_spacer_inner)
        control_layout.addWidget(self.btn_stop)
        control_layout.addSpacerItem(control_spacer_outer)

        moving_layout = QHBoxLayout()
        moving_layout.addSpacerItem(control_spacer_outer)
        moving_layout.addWidget(self.btn_move_back)
        moving_layout.addSpacerItem(control_spacer_inner)
        moving_layout.addWidget(self.btn_move_forw)
        moving_layout.addSpacerItem(control_spacer_outer)

        # buttons_layout.addWidget(self.btn_move_back)
        # buttons_layout.addWidget(self.btn_move_forw)
        left_vert_outer_spacer = QSpacerItem(300, 12)
        left_vert_inner_spacer = QSpacerItem(300, 32, QSizePolicy.Expanding, QSizePolicy.Fixed)

        left_scroll_layout = QVBoxLayout()
        left_scroll_layout.setContentsMargins(12, 16, 12, 12)
        left_scroll_layout.addWidget(self.scroll_area_left)
        self.scroll_group_left.setLayout(left_scroll_layout)

        left_layout = QVBoxLayout()
        left_layout.addItem(left_vert_outer_spacer)
        left_layout.addLayout(status_layout)

        batch_button_layout = QHBoxLayout()
        batch_button_layout.addSpacerItem(QSpacerItem(16, 32, QSizePolicy.Expanding, QSizePolicy.Minimum))
        batch_button_layout.addWidget(self.btn_new_batch)
        batch_button_layout.addSpacerItem(QSpacerItem(16, 32, QSizePolicy.Expanding, QSizePolicy.Minimum))
        left_layout.addLayout(batch_button_layout)

        left_layout.addLayout(control_layout)
        left_layout.addWidget(self.move_text)
        left_layout.addLayout(moving_layout)
        left_layout.addSpacerItem(left_vert_inner_spacer)
        left_layout.addWidget(self.scroll_group_left)
        left_layout.addItem(left_vert_outer_spacer)
        left_layout.setSpacing(12)

        self.btn_generate_report = QPushButton("Создать отчет", self)
        self.btn_generate_report.setFixedSize(240, 40)
        self.btn_generate_report.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 16px;
                border: 2px solid #000000;
                border-radius: 12px;
                padding: 8px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)

        self.btn_generate_report.clicked.connect(self.generate_report)

        # grbox = QHBoxLayout()
        # grbox.addStretch()
        # grbox.addWidget(self.btn_generate_report)
        # grbox.addStretch()
        # left_layout.addLayout(grbox)

        nav_layout = QHBoxLayout()
        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.btn_next)

        self.squares_widget = SquaresWidget(self.cam_manager)

        center_layout = QVBoxLayout()
        center_layout.addLayout(nav_layout)
        center_layout.addWidget(self.image_label)








        # Стекло (код 4)
        glass_label = QLabel("Стекло")
        glass_label.setFixedSize(80, 30)
        glass_label.setAlignment(Qt.AlignCenter)
        glass_label.setStyleSheet(f"""
            background-color: rgb({STATE_COLORS[4].red()}, {STATE_COLORS[4].green()}, {STATE_COLORS[4].blue()});
            color: white;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 12px;
            border-radius: 5px;
        """)

        # Сварка (код 8)
        weld_label = QLabel("Сварка")
        weld_label.setFixedSize(80, 30)
        weld_label.setAlignment(Qt.AlignCenter)
        weld_label.setStyleSheet(f"""
            background-color: rgb({STATE_COLORS[8].red()}, {STATE_COLORS[8].green()}, {STATE_COLORS[8].blue()});
            color: white;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 12px;
            border-radius: 5px;
        """)

        # Разновысотность (код 2)
        uneven_label = QLabel("Разновысотность")
        uneven_label.setFixedSize(120, 30)
        uneven_label.setAlignment(Qt.AlignCenter)
        uneven_label.setStyleSheet(f"""
            background-color: rgb({STATE_COLORS[2].red()}, {STATE_COLORS[2].green()}, {STATE_COLORS[2].blue()});
            color: white;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 12px;
            border-radius: 5px;
        """)

        # Раковины (код 3)
        sinks_label = QLabel("Раковины")
        sinks_label.setFixedSize(80, 30)
        sinks_label.setAlignment(Qt.AlignCenter)
        sinks_label.setStyleSheet(f"""
            background-color: rgb({STATE_COLORS[3].red()}, {STATE_COLORS[3].green()}, {STATE_COLORS[3].blue()});
            color: black;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 12px;
            border-radius: 5px;
        """)

        hbox = QHBoxLayout()
        hbox.addStretch()
        hbox.addWidget(glass_label)
        hbox.setSpacing(12)
        hbox.addWidget(sinks_label)
        hbox.setSpacing(12)
        hbox.addWidget(self.squares_widget)
        hbox.setSpacing(12)
        hbox.addWidget(uneven_label)
        hbox.setSpacing(12)
        hbox.addWidget(weld_label)
        hbox.addStretch()
        center_layout.addLayout(hbox)

        stats_group = QGroupBox("Статистика по изделиям")
        self.stats_group = stats_group
        stats_group.setFixedSize(350, 284)
        stats_group.setStyleSheet("""
            QGroupBox {
                font-family: "Montserrat";
                font-weight: bold;
                font-size: 20px;
                color: #343e43;
                background-color: #f5f5f5;
                border: 4px solid #d9d9d9;
                border-radius: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                top: -4px;
                padding: 0 16px;
                background-color: #f5f5f5;
            }
        """)

        stats_layout = QVBoxLayout()
        stats_layout.setSpacing(12)
        stats_layout.setContentsMargins(12, 32, 12, 12)

        self.good_label = QLabel("Годные изделия")
        self.bad_label = QLabel("Бракованные изделия")
        self.checked_label = QLabel("Всего проверено")

        for label in [self.good_label, self.bad_label, self.checked_label]:
            label.setFixedSize(160, 36)
            label.setStyleSheet("""
                QLabel {
                    font-family: "Montserrat";
                    font-size: 18px;
                    font-weight: bold;
                    color: #343e43;
                }
            """)

        self.good_amount = QLabel(str(self.squares_widget.stats["total_good"]))
        self.bad_amount = QLabel(str(self.squares_widget.stats["total_bad"]))
        self.checked_amount = QLabel(str(self.squares_widget.stats["total_checked"]))

        for amount in [self.good_amount, self.bad_amount, self.checked_amount]:
            amount.setFixedSize(70, 36)
            amount.setAlignment(Qt.AlignmentFlag.AlignCenter)
            amount.setStyleSheet("""
                QLabel {
                    background-color: #d9d9d9;
                    border: 2px solid #343e43;
                    border-radius: 12px;
                    font-family: "Montserrat";
                    font-size: 18px;
                    font-weight: bold;
                    color: #343e43;
                    padding: 8px 12px;
                }
            """)

        good_layout = QHBoxLayout()
        bad_layout = QHBoxLayout()
        checked_layout = QHBoxLayout()

        for layout, label, amount in [
            (good_layout, self.good_label, self.good_amount),
            (bad_layout, self.bad_label, self.bad_amount),
            (checked_layout, self.checked_label, self.checked_amount)]:
            layout.addWidget(label)
            layout.addItem(status_spacer_outer)
            layout.addStretch()
            layout.addWidget(amount)

        self.squares_widget.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 20px;
                border: 2px solid #000000;
                border-radius: 12px;
                padding: 8px;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;
            }
            QPushButton:hover {
                background-color: #8C1212;
                color: #FFFFFF;
            }
        """)

        self.batch_info_label = QLabel("Партия не начата")
        self.batch_info_label.setStyleSheet("""
            QLabel {
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 16px;
                background-color: #E8F5E9;
                border: 2px solid #4CAF50;
                border-radius: 8px;
                padding: 8px;
            }
        """)
        self.batch_info_label.setAlignment(Qt.AlignCenter)

        left_layout.insertWidget(3, self.batch_info_label)

        stats_layout.addLayout(good_layout)
        stats_layout.addLayout(bad_layout)
        stats_layout.addLayout(checked_layout)
        stats_layout.addSpacing(10)
        stats_layout.addWidget(self.squares_widget.clear_button)

        stats_group.setLayout(stats_layout)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(20)
        buttons_layout.addStretch()

        self.btn_toggle_thresholds = QPushButton("Пороговые значения", self)
        self.btn_toggle_thresholds.setFixedSize(240, 40)
        self.btn_toggle_thresholds.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 16px;
                border: 2px solid #000000;
                border-radius: 12px;
                padding: 8px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)
        self.btn_toggle_thresholds.clicked.connect(self.open_thresholds_dialog)

        self.btn_center = QPushButton("Центрировать деталь")
        self.btn_center.setFixedSize(240, 40)
        self.btn_center.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 16px;
                border: 2px solid #000000;
                border-radius: 16px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)
        self.btn_center.clicked.connect(self.center_part)

        buttons_layout.addWidget(self.btn_toggle_thresholds)
        buttons_layout.addWidget(self.btn_generate_report)
        # buttons_layout.addWidget(self.btn_move_back)
        # buttons_layout.addWidget(self.btn_move_forw)
        # buttons_layout.addWidget(self.btn_center)


        self.btn_toggle_stats = QPushButton("Статистика по изделиям", self)
        self.btn_toggle_stats.setFixedSize(240, 40)
        self.btn_toggle_stats.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 16px;
                border: 2px solid #000000;
                border-radius: 12px;
                padding: 8px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)
        self.btn_toggle_stats.clicked.connect(self.open_stats_dialog)
        buttons_layout.addWidget(self.btn_toggle_stats)

        buttons_layout.addStretch()
        center_layout.addLayout(buttons_layout)

        main_hor_spacer = QSpacerItem(40, 800, QSizePolicy.Fixed, QSizePolicy.Expanding)

        main_h_layout = QHBoxLayout()
        main_h_layout.addLayout(left_layout)
        main_h_layout.addItem(main_hor_spacer)
        main_h_layout.addLayout(center_layout)
        main_h_layout.addItem(main_hor_spacer)

        tab_widget = QTabWidget()

        tab1 = QWidget()
        tab1.setLayout(main_h_layout)
        tab_widget.addTab(tab1, "Основная")

        tab2 = QWidget()
        tab2_layout = QVBoxLayout(tab2)

        all_real_cameras = all(isinstance(cid, int) for cid in self.cam_manager.cam_ids)

        if all_real_cameras:
            self.live_manager = LiveCameraManager(self.cam_manager)
            self.live_label = QLabel()
            self.live_label.setFixedSize(1280, 720)
            self.live_label.setAlignment(Qt.AlignCenter)
            self.live_label.setStyleSheet("""
                QLabel {
                    background-color: #E0E0E0;
                    border-radius: 16px;
                }
            """)

            live_buttons_layout = QHBoxLayout()
            self.live_btn_prev = QPushButton("← Предыдущая камера")
            self.live_btn_next = QPushButton("Следующая камера →")
            for btn in (self.live_btn_prev, self.live_btn_next):
                btn.setFixedSize(340, 56)
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #D9D9D9;
                        color: #343E43;
                        font-family: Montserrat;
                        font-weight: bold;
                        font-size: 14px;
                        border: 2px solid #000000;
                        border-radius: 12px;
                        padding: 8px;
                    }
                    QPushButton:hover {
                        background-color: #3997E4;
                        color: #FFFFFF;
                    }
                """)
            live_buttons_layout.addStretch()
            live_buttons_layout.addWidget(self.live_btn_prev)
            live_buttons_layout.addWidget(self.live_btn_next)
            live_buttons_layout.addStretch()

            live_control_layout = QHBoxLayout()
            self.live_btn_start = QPushButton("Начать проверку")
            self.live_btn_stop = QPushButton("Стоп")
            for btn in (self.live_btn_start, self.live_btn_stop):
                btn.setFixedSize(200, 56)
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #D9D9D9;
                        color: #343E43;
                        font-family: Montserrat;
                        font-weight: bold;
                        font-size: 16px;
                        border: 2px solid #000000;
                        border-radius: 16px;
                        padding: 8px;
                    }
                    QPushButton:hover {
                        background-color: #3997E4;
                        color: #FFFFFF;
                    }
                """)
            self.live_btn_stop.setEnabled(False)


            # Layout для управления движением
            move_layout = QHBoxLayout()
            move_layout.setSpacing(10)
            move_layout.addStretch()

            # Кнопка "Назад" (стрелка влево)
            self.btn_move_backward = QPushButton("←")
            self.btn_move_backward.setFixedSize(60, 60)
            self.btn_move_backward.setStyleSheet("""
                QPushButton {
                    background-color: #D9D9D9;
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 24px;
                    border: 2px solid #000000;
                    border-radius: 12px;
                }
                QPushButton:hover {
                    background-color: #3997E4;
                    color: #FFFFFF;
                }
            """)
            move_layout.addWidget(self.btn_move_backward)

            # Поле ввода
            self.move_distance_edit = QLineEdit()
            self.move_distance_edit.setFixedSize(100, 60)
            self.move_distance_edit.setAlignment(Qt.AlignCenter)
            self.move_distance_edit.setPlaceholderText(" ")
            self.move_distance_edit.setValidator(QIntValidator())
            self.move_distance_edit.setStyleSheet("""
                QLineEdit {
                    background-color: white;
                    border: 2px solid #000000;
                    border-radius: 12px;
                    font-family: Montserrat;
                    font-size: 18px;
                    font-weight: bold;
                    color: #343E43;
                    padding: 8px;
                }
                QLineEdit:focus {
                    border: 2px solid #3997E4;
                }
            """)
            move_layout.addWidget(self.move_distance_edit)

            # Комбобокс выбора режима
            self.move_mode_combo = QComboBox()
            self.move_mode_combo.addItems(["0.1 мм", "1 мм", "1 см", "инженерный"])
            self.move_mode_combo.setFixedSize(100, 40)
            self.move_mode_combo.setStyleSheet("""
                QComboBox {
                    background-color: #D9D9D9;
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 14px;
                    border: 2px solid #000000;
                    border-radius: 8px;
                    padding: 5px;
                }
                QComboBox:hover {
                    background-color: #3997E4;
                    color: #FFFFFF;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox QAbstractItemView {
                    background-color: #f5f5f5;
                    color: #343E43;
                    selection-background-color: #3997E4;
                    selection-color: #FFFFFF;
                }
            """)
            self.move_mode_combo.currentIndexChanged.connect(self.on_move_mode_changed)
            move_layout.addWidget(self.move_mode_combo)

            # Кнопка "Вперёд" (стрелка вправо)
            self.btn_move_forward = QPushButton("→")
            self.btn_move_forward.setFixedSize(60, 60)
            self.btn_move_forward.setStyleSheet("""
                QPushButton {
                    background-color: #D9D9D9;
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 24px;
                    border: 2px solid #000000;
                    border-radius: 12px;
                }
                QPushButton:hover {
                    background-color: #3997E4;
                    color: #FFFFFF;
                }
            """)
            move_layout.addWidget(self.btn_move_forward)
            move_layout.addStretch()

            live_control_layout.addStretch()
            live_control_layout.addWidget(self.live_btn_start)
            live_control_layout.addWidget(self.live_btn_stop)
            # live_control_layout.addWidget(self.btn_move_forw)
            # live_control_layout.addWidget(self.btn_move_back)
            live_control_layout.addStretch()

            # Инициализация множителя
            self.move_multiplier = 1
            self.on_move_mode_changed(0)  # установим начальное значение

            # Добавляем в tab2_layout после создания move_layout
            tab2_layout.addLayout(move_layout)
            tab2_layout.addStretch()
            tab2_layout.addWidget(self.live_label, alignment=Qt.AlignHCenter)
            tab2_layout.addLayout(live_buttons_layout)
            tab2_layout.addLayout(live_control_layout)
            tab2_layout.addStretch()

            self.live_current_cam = 0
            self.live_timer = QTimer()
            self.live_timer.timeout.connect(self.update_live_ui)

            self.live_btn_prev.clicked.connect(self.live_prev_camera)
            self.live_btn_next.clicked.connect(self.live_next_camera)
            self.live_btn_start.clicked.connect(self.start_live)
            self.live_btn_stop.clicked.connect(self.stop_live)
            self.btn_move_backward.clicked.connect(self.move_backward)
            self.btn_move_forward.clicked.connect(self.move_forward)

        else:
            self.live_manager = None
            label_unavailable = QLabel("Прямой эфир недоступен в тестовом режиме")
            label_unavailable.setAlignment(Qt.AlignCenter)
            label_unavailable.setStyleSheet("font-size: 20px; color: gray;")
            tab2_layout.addWidget(label_unavailable)

        tab_widget.addTab(tab2, "Прямой эфир")

        tab3 = QWidget()
        # Главный вертикальный макет для вкладки
        tab3_layout = QVBoxLayout(tab3)
        tab3_layout.setContentsMargins(10, 10, 10, 10)  # Небольшие отступы от краев вкладки

        # ==========================================
        # 1. ВЕРХНЯЯ ЧАСТЬ: Только кнопки и ввод
        # ==========================================
        top_section_layout = QHBoxLayout()
        top_section_layout.setAlignment(Qt.AlignCenter)

        # Поле ввода номера партии
        self.batch_number_text = QLabel("Номер партии:")
        self.batch_number_text.setStyleSheet("""
            color: #343E43;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 24px;
        """)

        self.batch_number_edit = QLineEdit()
        self.batch_number_edit.setFixedSize(100, 60)
        self.batch_number_edit.setAlignment(Qt.AlignCenter)
        self.batch_number_edit.setPlaceholderText("№")
        self.batch_number_edit.setValidator(QIntValidator())
        self.batch_number_edit.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 2px solid #000000;
                border-radius: 12px;
                font-family: Montserrat;
                font-size: 18px;
                font-weight: bold;
                color: #343E43;
                padding: 8px;
            }
            QLineEdit:focus {
                border: 2px solid #3997E4;
            }
        """)

        # Кнопка загрузки партии по номеру
        btn_load_by_number = QPushButton("Загрузить партию")
        btn_load_by_number.clicked.connect(self.load_batch_by_number)
        btn_load_by_number.setFixedSize(300, 60)
        btn_load_by_number.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 12px;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)

        # Кнопка последней партии
        btn_last_completed = QPushButton("Последняя партия")
        btn_last_completed.clicked.connect(self.load_last_completed_batch)
        btn_last_completed.setFixedSize(300, 60)
        btn_last_completed.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 12px;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)

        # Кнопка отчёта (исправлена ошибка копирования setFixedSize)
        self.btn_open_report = QPushButton("Открыть отчёт")
        self.btn_open_report.setFixedSize(300, 60)  # Было ошибочно присвоено btn_last_completed
        self.btn_open_report.setEnabled(False)
        self.btn_open_report.setStyleSheet("""
            QPushButton {
                background-color: #D9D9D9;
                color: #343E43;
                font-family: Montserrat;
                font-weight: bold;
                font-size: 24px;
                border: 2px solid #000000;
                border-radius: 12px;
                padding: 12px;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;
            }
            QPushButton:hover {
                background-color: #3997E4;
                color: #FFFFFF;
            }
        """)
        self.btn_open_report.clicked.connect(self.open_current_report)

        # Добавляем все элементы в верхний горизонтальный макет
        top_section_layout.addWidget(self.batch_number_text)
        top_section_layout.addWidget(self.batch_number_edit)
        top_section_layout.addWidget(btn_load_by_number)
        top_section_layout.addWidget(btn_last_completed)
        top_section_layout.addWidget(self.btn_open_report)
        top_section_layout.addStretch()

        # Добавляем верхний блок в главный макет
        tab3_layout.addLayout(top_section_layout)

        # ==========================================
        # 2. НИЖНЯЯ ЧАСТЬ: График (слева) и Метка (справа)
        # ==========================================
        bottom_section_layout = QHBoxLayout()

        # ЛЕВАЯ ЧАСТЬ: График
        self.figure = Figure(figsize=(6, 4))
        self.canvas = FigureCanvas(self.figure)
        # Лучше не использовать setFixedSize для графика, чтобы он растягивался
        # Но если нужно строго:
        self.figure.set_facecolor('#f5f5f5')
        self.canvas.setFixedSize(500, 800)
        # stretch=4 означает, что график займет 4 части пространства (шире)
        bottom_section_layout.addWidget(self.canvas, stretch=4)

        # ПРАВАЯ ЧАСТЬ: Метка статистики
        self.batch_stats_label = QLabel("Статистика партии")
        self.batch_stats_label.setStyleSheet("""
            color: #343E43;
            font-family: Montserrat;
            font-weight: bold;
            font-size: 24px;
            background-color: #ffffff;
            border: 2px solid #D9D9D9;
            border-radius: 12px;
            padding: 20px;
        """)
        self.batch_stats_label.setAlignment(Qt.AlignCenter)
        self.batch_stats_label.setWordWrap(True)  # Перенос длинного текста
        # stretch=1 означает, что метка займет 1 часть пространства (уже)
        bottom_section_layout.addWidget(self.batch_stats_label, stretch=1)

        # Добавляем нижний блок в главный макет
        tab3_layout.addLayout(bottom_section_layout)

        # Добавляем вкладку в таб-виджет
        tab_widget.addTab(tab3, "Статистика по партиям")

        main_layout = QVBoxLayout()
        main_layout.addWidget(tab_widget)
        self.setLayout(main_layout)


        self.check_and_restore_states()

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_ui)
        self.timer.start(5)
        self.update_ui()
        self.showMaximized()
        self.check_and_display_active_batch()

        self.stats_updater = StatsUpdater(self.cam_manager)
        self.stats_updater.stats_updated.connect(self.on_stats_updated)

        self.stats_timer = QTimer()
        self.stats_timer.timeout.connect(self.update_statistics)
        self.stats_timer.start(100)

        self.batch_timer = QTimer()
        self.batch_timer.timeout.connect(self.check_batch_status)
        self.batch_timer.start(500)

        self.bool_switch_ways = True

    # ---------- Методы для работы с партиями и состоянием ----------
    def check_and_display_active_batch(self):
        if self.cam_manager.is_batch_active():
            batch_info = self.cam_manager.get_current_batch_info()
            if batch_info:
                self.batch_info_label.setText(
                    f"Активная партия:\n"
                    f"{batch_info['batch_name']}\n"
                    f"Начало: {batch_info['start_time']}"
                )
                print(f"Восстановлена активная партия: {batch_info['batch_name']}")
                self.btn_new_batch.setText("Партия активна")
                self.btn_new_batch.setEnabled(False)
            else:
                self.batch_info_label.setText("Ошибка загрузки информации о партии")
        else:
            self.batch_info_label.setText("Партия не активна")
            print("Активная партия не найдена")

    def start_new_batch(self):
        """Начинает новую партию сохранения скриншотов."""
        try:
            print("Нажата кнопка 'Начать новую партию'")

            # Если уже есть активная партия, завершаем ее
            if self.cam_manager.is_batch_active():
                # Спрашиваем подтверждение
                from PyQt5.QtWidgets import QMessageBox
                reply = QMessageBox.question(
                    self, 'Подтверждение',
                    'Уже есть активная партия. Начать новую партию?',
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )

                if reply == QMessageBox.No:
                    return

            # Начинаем новую партию
            # dialog = BatchNameDialog(self)
            # if dialog.exec_() != QDialog.Accepted:
            #     return
            # custom_name = dialog.get_name()
            # batch_path = self.cam_manager.start_new_batch(custom_name)


            batch_path = self.cam_manager.start_new_batch()
            if batch_path:
                print(f"Партия создана успешно: {batch_path}")
                batch_info = self.cam_manager.get_current_batch_info()
                if batch_info:
                    self.batch_info_label.setText(
                        f"Активная партия:\n"
                        f"{batch_info['batch_name']}\n"
                        f"Начало: {batch_info['start_time']}"
                    )
                    print(f"Информация о партии: {batch_info['batch_name']}")
                else:
                    self.batch_info_label.setText("Партия начата, но информация недоступна")
                    print("Информация о партии недоступна")
                self.btn_new_batch.setEnabled(False)
                self.btn_new_batch.setText("Партия активна")
            else:
                print("Не удалось создать партию (batch_path = None)")
                QMessageBox.warning(self, "Ошибка", "Не удалось создать новую партию")
        except Exception as e:
            print(f"Исключение при начале новой партии: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Критическая ошибка",
                                 f"Не удалось начать новую партию:\n{str(e)}")

    def start_conveyor(self):
        self.stop_live()
        self.cam_manager.start_processing()
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_center.setEnabled(False)  # <-- Было btn_center (вызывало NameError)
        self.btn_new_batch.setEnabled(False)
        if hasattr(self, 'live_btn_start') and self.live_btn_start:
            self.live_btn_start.setEnabled(False)
        self.status_indicator.setStyleSheet("background-color: #4CAF50; border-radius: 16px; ")
        self.status_text.setText("Система работает")
        self.cam_manager.safe_send_command("G8 S1")
        self.cam_manager.safe_send_command("G7 S18803")
        if self.cam_manager.is_batch_active():
            batch_info = self.cam_manager.get_current_batch_info()
            if batch_info:
                self.batch_info_label.setText(
                    f"Активная партия:\n"
                    f"{batch_info['batch_name']}\n"
                    f"Начало: {batch_info['start_time']}\n"
                    f"Статус: Запись"
                )

    def stop_conveyor(self):
        self.cam_manager.stop_processing()
        self.start_live()
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_center.setEnabled(True)  # <-- Было btn_center (вызывало NameError)
        if self.cam_manager.is_batch_active():
            self.btn_new_batch.setEnabled(False)
        else:
            self.btn_new_batch.setEnabled(True)
            self.btn_new_batch.setText("Начать новую партию")
        if hasattr(self, 'live_btn_start') and self.live_btn_start:
            self.live_btn_start.setEnabled(True)
        self.status_indicator.setStyleSheet("background-color: #8C1212; border-radius: 16px; ")
        self.status_text.setText("Система остановлена")
        if self.cam_manager.is_batch_active():
            batch_info = self.cam_manager.get_current_batch_info()
            if batch_info:
                self.batch_info_label.setText(
                    f"Активная партия:\n"
                    f"{batch_info['batch_name']}\n"
                    f"Начало: {batch_info['start_time']}\n"
                    f"Статус: Остановлена"
                )

    def start_live(self):
        if self.live_manager:
            self.live_manager.start()
            self.live_timer.start(30)
            self.live_btn_start.setEnabled(False)
            self.live_btn_stop.setEnabled(True)
            if hasattr(self, 'btn_start'):
                self.btn_start.setEnabled(False)

    def stop_live(self):
        if self.live_manager:
            self.live_timer.stop()
            self.live_manager.stop()
            self.live_btn_start.setEnabled(True)
            self.live_btn_stop.setEnabled(False)
            if hasattr(self, 'btn_start'):
                self.btn_start.setEnabled(True)
            self.live_label.clear()
            self.live_label.setText("Прямой эфир остановлен")
            self.live_label.setAlignment(Qt.AlignCenter)
            self.cam_manager.safe_send_command("G8 S1")
            self.cam_manager.safe_send_command("G7 S18803")

    # ---------- Движение ----------

    def kostil(self):
        try:
            self.cam_manager.safe_send_command("G7 S1")
            self.cam_manager.safe_send_command("G3")

            # Опрос статуса
            while True:
                self.cam_manager.safe_send_command("I1")
                # Ждём ответ с таймаутом
                response = self.cam_manager.ser.readline().strip().decode()
                if response == '0':
                    break
                elif response == '1':
                    # Движение выполняется
                    pass
                else:
                    print(f"{response}")
                # Обрабатываем события и ждём 0.1 сек 10 раз = 1 сек
                for _ in range(10):
                    QApplication.processEvents()
                    time.sleep(0.1)

            # Возврат в исходное состояние
            print(f"костыль сработал")
        except Exception as e:
            print(f"Ошибка костыля: {e}")

    def on_move_mode_changed(self, index):
        """Обработчик изменения режима движения."""
        mode = self.move_mode_combo.currentText()
        if mode == "0.1 мм":
            self.move_multiplier = 35
        elif mode == "1 мм":
            self.move_multiplier = 346
        elif mode == "1 см":
            self.move_multiplier = 3463
        else:  # инженерный
            self.move_multiplier = 1
        print(f"Режим движения: {mode}, множитель: {self.move_multiplier}")

    def move_backward(self):
        """Движение назад (бывшее влево)."""
        # Блокируем элементы управления
        self.move_distance_edit.setEnabled(False)
        self.btn_move_forward.setEnabled(False)
        self.btn_move_backward.setEnabled(False)

        distance = self.get_move_distance() * self.move_multiplier
        if distance <= 0:
            print("Расстояние должно быть положительным")
            self.move_distance_edit.setEnabled(True)
            self.btn_move_forward.setEnabled(True)
            self.btn_move_backward.setEnabled(True)
            return
        try:
            self.cam_manager.safe_send_command("G8 S0")
            if not self.bool_switch_ways:
                self.kostil()
                self.bool_switch_ways = True
            self.cam_manager.safe_send_command(f"G7 S{distance}")
            self.cam_manager.safe_send_command("G3")

            # Возврат в исходное состояние
            print(f"Движение назад на {distance} завершено")
        except Exception as e:
            print(f"Ошибка движения назад: {e}")
        finally:
            # Разблокируем элементы
            self.move_distance_edit.setEnabled(True)
            self.btn_move_forward.setEnabled(True)
            self.btn_move_backward.setEnabled(True)

    def move_forward(self):
        """Движение вперёд (бывшее правое)."""
        self.move_distance_edit.setEnabled(False)
        self.btn_move_forward.setEnabled(False)
        self.btn_move_backward.setEnabled(False)

        distance = self.get_move_distance() * self.move_multiplier
        if distance <= 0:
            print("Расстояние должно быть положительным")
            self.move_distance_edit.setEnabled(True)
            self.btn_move_forward.setEnabled(True)
            self.btn_move_backward.setEnabled(True)
            return
        try:
            self.cam_manager.safe_send_command("G8 S1")
            if not self.bool_switch_ways:
                self.kostil()
                self.bool_switch_ways = True
            self.cam_manager.safe_send_command(f"G7 S{distance}")
            self.cam_manager.safe_send_command("G3")

            print(f"Движение вперёд на {distance} завершено")
        except Exception as e:
            print(f"Ошибка движения вперёд: {e}")
        finally:
            self.move_distance_edit.setEnabled(True)
            self.btn_move_forward.setEnabled(True)
            self.btn_move_backward.setEnabled(True)

    def center_part(self):
        counter = 6
        temp = 0
        self.btn_start.setEnabled(False)
        self.start_live()
        time.sleep(3)
        if not self.live_manager:
            QMessageBox.warning(self, "Ошибка", "Прямой эфир недоступен")
            return
        # Получаем текущий кадр с выбранной камеры
        try:
            while counter > 0 and temp == 0:
                frame = self.live_manager.get_frame(1)
                if frame is None:
                    QMessageBox.warning(self, "Ошибка", "Не удалось получить кадр с камеры")
                    return
                # Сохраняем кадр во временный файл
                temp_path = Path(f"{counter}temp_center.jpg")
                cv2.imwrite(str(temp_path), frame)
                # try:
                # Вызываем функцию центрирования
                temp = draw_part_contour(
                    cam_manager=self.cam_manager,
                    image_path=temp_path,
                    model_path="new_weights/backside.pt",  # укажите путь к вашей сегментационной модели
                )
                counter -= 1
            QMessageBox.information(self, "Успех", "Центрирование выполнено")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка центрирования: {str(e)}")
        finally:
            self.btn_start.setEnabled(True)
            self.stop_live()
            counter = 6
            temp = 0
            # if temp_path.exists():
            #     temp_path.unlink()

    def get_move_distance(self):
        text = self.move_distance_edit.text().strip()
        if text:
            try:
                return int(text)
            except ValueError:
                return 0
        return 0

    def on_pressed_forw(self):
        self.cam_manager.safe_send_command("G7 S100000")
        self.cam_manager.safe_send_command("G3")

    def on_released_forw(self):
        self.cam_manager.safe_send_command("G1")

    def on_pressed_back(self):
        self.cam_manager.safe_send_command("G8 S0")
        self.cam_manager.safe_send_command("G7 S100000")
        self.cam_manager.safe_send_command("G3")


    def on_released_back(self):
        self.cam_manager.safe_send_command("G1")
        self.cam_manager.safe_send_command("G8 S1")
        self.cam_manager.safe_send_command("G7 S50")
        self.cam_manager.safe_send_command("G3")


    # ---------- Обновление прямого эфира ----------
    def update_live_ui(self):
        if self.live_manager:
            frame = self.live_manager.get_frame(self.live_current_cam)
            if frame is not None:
                h, w, ch = frame.shape
                bytes_per_line = ch * w
                qimg = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(qimg)
                self.live_label.setPixmap(pixmap)
            else:
                self.live_label.setText("Нет сигнала с камеры")
                self.live_label.setAlignment(Qt.AlignCenter)

    def live_prev_camera(self):
        if self.live_manager:
            self.live_current_cam = (self.live_current_cam - 1) % len(self.cam_manager.cam_ids)

    def live_next_camera(self):
        if self.live_manager:
            self.live_current_cam = (self.live_current_cam + 1) % len(self.cam_manager.cam_ids)

    # ---------- Статистика и обновление UI ----------
    def on_stats_updated(self, stats):
        self.good_amount.setText(str(stats["total_good"]))
        self.bad_amount.setText(str(stats["total_bad"]))
        self.checked_amount.setText(str(stats["total_checked"]))
        self.squares_widget.stats = stats
        self.squares_widget.update()

    def update_statistics(self):
        self.stats_updater.update_stats()

    def update_ui(self):
        frame = self.cam_manager.get_frame(self.current_cam_index)
        self.btn_generate_report.setEnabled(self.cam_manager.is_batch_active())
        if frame is not None:
            h, w, ch = frame.shape
            bytes_per_line = ch * w
            qimg = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qimg)
            self.image_label.setPixmap(pixmap)
            self.update_object_info(frame)
        else:
            self.image_label.setMovie(self.anim)
            self.info_table.setText("Нет данных об объектах")
        self.squares_widget.update()
        is_running = self.cam_manager.running
        self.btn_new_batch.setEnabled(not is_running)
        if is_running and self.cam_manager.is_batch_active():
            self.btn_new_batch.setText("Партия активна (работа)")
            self.btn_new_batch.setStyleSheet("""
                QPushButton {
                    background-color: #D9D9D9;
                    color: #343E43;
                    font-family: Montserrat;
                    font-weight: bold;
                    font-size: 24px;
                    border: none;
                    border-radius: 16px;
                    padding: 8px;
                }
                QPushButton:disabled {
                    background-color: #B0B0B0;
                }
                QPushButton:hover {
                    background-color: #3997E4;
                    color: #FFFFFF;
                }
            """)

    def check_batch_status(self):
        if self.cam_manager.is_batch_active():
            batch_info = self.cam_manager.get_current_batch_info()
            if batch_info:
                try:
                    batch_path = self.cam_manager.batch_manager.current_batch_path
                    if batch_path:
                        stats_file = batch_path / "batch_statistics.json"
                        if stats_file.exists():
                            import json
                            with open(stats_file, 'r', encoding='utf-8') as f:
                                batch_stats = json.load(f)
                except Exception as e:
                    print(f"Ошибка загрузки статистики партии: {e}")

    def check_and_restore_states(self):
        states = self.cam_manager.load_initial_states()
        if states and any(v != 0 for v in states):
            reply = QMessageBox.question(
                self,
                "Восстановление состояний",
                "При выключении программы на конвейере оставались проверенные, но нераспределенные детали.\n"
                "Были ли они убраны с ленты?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.cam_manager.clear_initial_states()
                QMessageBox.information(self, "Очищено", "Файл состояний очищен.")
            else:
                self.cam_manager.restore_states(states)
                QMessageBox.information(self, "Восстановлено",
                                        f"Состояния первых четырёх позиций восстановлены:\n{states}")

    def update_object_info(self, frame):
        object_data = self.cam_manager.get_object_data(self.current_cam_index)
        objects_by_type = {}
        for obj in object_data:
            class_name = obj["class_name"]
            objects_by_type.setdefault(class_name, []).append(obj)

        name_map = {
            "Omission": "Пропуск",
            "omission": "Пропуск",
            "flatness": "Плоскостность",
            "glass": "Стекло",
            "sinks": "Провал",
            "platform": "Платформа",
            "contacts": "Контакт",
            "contacts_long": "Длинный контакт",
            "flatness_short": "Короткая плоскостность",
            "window_sinks": "Раковина в окне",
            "objects": "Объекты",
            "mechanics": "Механика"
        }
        camera_names = {
            0: "Камера 1 (Центральная)",
            1: "Камера 2 (Ближняя)",
            2: "Камера 3 (Дальняя)"
        }
        camera_name = camera_names.get(self.current_cam_index, f"Камера {self.current_cam_index + 1}")

        html = f"""
        <html>
        <head>
        <style>
            body {{ margin:0; padding:0; font-family:Arial; }}
            .camera-name {{ font-size:14px; font-weight:bold; margin-bottom:10px; color:#333; }}
            table {{ border-collapse:collapse; width:100%; }}
            th, td {{ border:1px solid #ddd; padding:5px; text-align:left; font-size:12px; }}
            th {{ background-color:#f2f2f2; position:sticky; top:0; }}
            .confidence {{ color:#666; font-size:0.9em; }}
            .object-type {{ font-weight:bold; background-color:#e6e6e6; }}
            .no-objects {{ color:#888; font-style:italic; padding:10px; }}
        </style>
        </head>
        <body>
        <div class="camera-name">{camera_name}</div>
        <table><thead><tr><th>Тип объекта</th><th>Параметры</th></tr></thead><tbody>
        """

        if not object_data:
            html += '<tr><td colspan="2" class="no-objects">Объекты не обнаружены</td></tr>'
        else:
            if "platform" in objects_by_type:
                platform = objects_by_type["platform"][0]
                segment = np.array(platform["segment"])
                x, y, w, h = cv2.boundingRect(segment)
                html += f"""
                <tr><td class="object-type">Платформа</td>
                <td>Ш: {w}px, В: {h}px<br>Площадь: {platform['area']:.0f}px²<br>
                <span class="confidence">Уверенность: {platform['confidence']*100:.1f}%</span></td></tr>
                """
            if "contacts" in objects_by_type:
                for i, contact in enumerate(sorted(objects_by_type["contacts"], key=lambda x: x.get("contact_id",0)),1):
                    html += f"""
                    <tr><td>Контакт {i}</td>
                    <td>Площадь: {contact['area']:.0f}px²<br>
                    <span class="confidence">Уверенность: {contact['confidence']*100:.1f}%</span></td></tr>
                    """
            for obj_type, obj_list in objects_by_type.items():
                if obj_type in ("platform","contacts"):
                    continue
                display_name = name_map.get(obj_type, obj_type)
                html += f'<tr><td colspan="2" class="object-type">{display_name} (всего: {len(obj_list)})</td></tr>'
                for i, obj in enumerate(obj_list,1):
                    additional = ""
                    if obj_type in ("Omission","omission") and "distance" in obj:
                        additional = f"<br>Расстояние: {obj['distance']:.1f}px"
                    html += f"""
                    <tr><td>{display_name} {i}</td>
                    <td>Площадь: {obj['area']:.0f}px²{additional}<br>
                    <span class="confidence">Уверенность: {obj['confidence']*100:.1f}%</span></td></tr>
                    """
        html += "</tbody></table></body></html>"
        self.info_table.setText(html)
        self.info_table.adjustSize()

    def prev_camera(self):
        self.current_cam_index = (self.current_cam_index - 1) % len(self.cam_manager.cam_ids)
        self.update_ui()

    def next_camera(self):
        self.current_cam_index = (self.current_cam_index + 1) % len(self.cam_manager.cam_ids)
        self.update_ui()

    def generate_report(self):
        if not self.cam_manager.is_batch_active():
            QMessageBox.warning(self, "Внимание", "Нет активной партии для создания отчёта.")
            return
        dialog = ReportDialog(self)
        if dialog.exec_() != QDialog.Accepted:
            return
        filename, count_str = dialog.get_values()
        if not filename:
            filename = "report"
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        max_items = None
        if count_str.strip():
            try:
                max_items = int(count_str)
                if max_items <= 0:
                    QMessageBox.warning(self, "Ошибка", "Количество должно быть положительным числом.")
                    return
            except ValueError:
                QMessageBox.warning(self, "Ошибка", "Количество должно быть целым числом.")
                return
        batch_info = self.cam_manager.get_current_batch_info()
        if not batch_info:
            QMessageBox.warning(self, "Ошибка", "Не удалось получить информацию о партии.")
            return
        batch_path = batch_info.get("batch_path")
        if not batch_path:
            QMessageBox.warning(self, "Ошибка", "Путь к партии не определён.")
            return
        output_file = Path(batch_path) / filename
        try:
            self.btn_generate_report.setEnabled(False)
            self.btn_generate_report.setText("Генерация...")
            QApplication.processEvents()
            result = generate_report(batch_path, output_path=output_file, max_items=max_items)
            if result:
                QMessageBox.information(self, "Успех", f"Отчёт успешно создан:\n{result}")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось создать отчёт.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при создании отчёта:\n{str(e)}")
        finally:
            self.btn_generate_report.setEnabled(True)
            self.btn_generate_report.setText("📁")

    def open_thresholds_dialog(self):
        dialog = ThresholdsDialog(self.cam_manager, self)
        dialog.exec_()

    def open_stats_dialog(self):
        dialog = StatsDialog(self.cam_manager, self)
        dialog.exec_()
    #
    # def closeEvent(self, event):
    #     if hasattr(self, 'live_manager') and self.live_manager:
    #         self.live_manager.release()
    #     self.cam_manager.save_conveyor_states()
    #     self.cam_manager.release()
    #     event.accept()

    def get_batch_folder_by_index(self, batch_index):
        """Возвращает путь к папке партии по её номеру (индексу)."""
        screenshots_dir = Path("utils/screenshots")
        if not screenshots_dir.exists():
            return None
        for folder in screenshots_dir.iterdir():
            if folder.is_dir() and folder.name.startswith(f"партия_{batch_index}_"):
                return folder
        return None

    def get_last_completed_batch_folder(self):
        """Возвращает путь к папке последней завершённой партии (имеющей final_report.xlsx)."""
        screenshots_dir = Path("utils/screenshots")
        if not screenshots_dir.exists():
            return None
        completed = []
        for folder in screenshots_dir.iterdir():
            if folder.is_dir() and folder.name.startswith("партия_"):
                report = folder / "final_report.xlsx"
                if report.exists():
                    # Извлекаем индекс из имени папки
                    try:
                        parts = folder.name.split('_')
                        if len(parts) >= 2 and parts[0] == "партия":
                            index = int(parts[1])
                            completed.append((index, folder))
                    except:
                        pass
        if not completed:
            return None
        completed.sort(key=lambda x: x[0], reverse=True)
        return completed[0][1]

    def parse_final_report_and_plot(self, batch_path):
        """Читает final_report.xlsx, строит гистограмму брака и круговую диаграмму качества."""
        report_path = batch_path / "final_report.xlsx"
        self.current_report_path = report_path
        self.btn_open_report.setEnabled(True)
        batch_num = batch_path.name.split('_')
        if not report_path.exists():
            QMessageBox.warning(self, "Ошибка", f"Файл отчёта не найден:\n{report_path}")
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(report_path)
            ws = wb.active

            # --- ДОБАВЛЕНО: чтение данных из ячеек ---
            workstation = ws['B3'].value if ws['B3'].value else 'N/A'
            product_type = ws['B4'].value if ws['B4'].value else 'N/A'
            batch_id = ws['B5'].value if ws['B5'].value else 'N/A'
            start_time = ws['B7'].value if ws['B7'].value else 'N/A'
            end_time = ws['B8'].value if ws['B8'].value else 'N/A'
            # ---------------------------------------

            # Поиск заголовка "Тип брака"
            header_row = 11
            defect_col = None
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
                for cell in row:
                    if cell.value == "Тип брака":
                        defect_col = cell.column
                        header_row = cell.row
                        break
                if defect_col:
                    break

            # Подсчёт типов брака
            defect_counts = {}
            if defect_col:
                for row in ws.iter_rows(min_row=header_row + 1, max_col=defect_col):
                    cell = row[defect_col - 1]
                    if cell.value:
                        types = [t.strip() for t in str(cell.value).split(',') if t.strip()]
                        for t in types:
                            defect_counts[t] = defect_counts.get(t, 0) + 1

            # Поиск столбца "Результат проверки"
            result_col = None
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
                for cell in row:
                    if cell.value == "Результат проверки":
                        result_col = cell.column
                        break
                if result_col:
                    break

            good_count = 0
            bad_count = 0
            clean_count = 0
            if result_col:
                for row in ws.iter_rows(min_row=header_row + 1, max_col=result_col):
                    cell = row[result_col - 1]
                    if cell.value == "брак":
                        bad_count += 1
                    elif cell.value == "чистка":
                        clean_count += 1
                    elif cell.value == "годен":
                        good_count += 1

            # Очистка и создание двух субплотов
            self.figure.clear()
            ax1 = self.figure.add_subplot(2, 1, 2)  # гистограмма
            ax2 = self.figure.add_subplot(2, 1, 1)  # круговая диаграмма

            # Гистограмма
            if defect_counts:
                labels = list(defect_counts.keys())
                values = list(defect_counts.values())
                bars = ax1.bar(labels, values, color='steelblue', edgecolor='black', linewidth=1.5, hatch='',
                               alpha=0.9)
                ax1.set_xlabel('Тип брака')
                ax1.set_ylabel('Количество')
                ax1.set_title('Статистика брака')
                ax1.set_xticklabels(labels, rotation=45, ha='right')
                ax1.grid(True, linestyle='--', alpha=0.7)
                for bar in bars:
                    height = bar.get_height()
                    ax1.annotate(f'{height}',
                                 xy=(bar.get_x() + bar.get_width() / 2, height),
                                 xytext=(0, 3),
                                 textcoords="offset points",
                                 ha='center', va='bottom')
            else:
                ax1.text(0.5, 0.5, 'Нет данных о браке', transform=ax1.transAxes, ha='center', va='center')
                ax1.set_title('Статистика брака')

            # Круговая диаграмма
            sizes = [good_count, bad_count, clean_count]
            labels_pie = ['Годные', 'Бракованные', 'Чистка']
            colors_pie = ['#66b3ff', '#ff9999', '#ffcc99']

            # Фильтрация нулевых значений
            filtered_sizes = []
            filtered_labels = []
            filtered_colors = []
            for size, label, color in zip(sizes, labels_pie, colors_pie):
                if size > 0:
                    filtered_sizes.append(size)
                    filtered_labels.append(label)
                    filtered_colors.append(color)

            if filtered_sizes:
                ax2.pie(filtered_sizes, labels=filtered_labels, colors=filtered_colors,
                        autopct='%1.1f%%', startangle=90)
                ax2.set_title('Распределение по качеству')
            else:
                ax2.text(0.5, 0.5, 'Нет данных о качестве', transform=ax2.transAxes,
                         ha='center', va='center')
                ax2.set_title('Распределение по качеству')
            ax2.axis('equal')

            self.figure.tight_layout()
            self.canvas.draw()

            # --- ИЗМЕНЕНО: вывод расширенной информации ---
            info_text = (f"Партия № {batch_num[1]}\n"
                         f"Рабочее место: {workstation}\n"
                         f"Тип изделия: {product_type}\n"
                         f"Идентификатор партии: {batch_id}\n"
                         f"Начало: {start_time}\n"
                         f"Конец: {end_time}")
            self.batch_stats_label.setText(info_text)
            # -------------------------------------------

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось обработать отчёт:\n{str(e)}")

    def load_batch_by_number(self):
        """Загружает партию по номеру, введённому в поле."""
        number_str = self.batch_number_edit.text().strip()
        if not number_str:
            QMessageBox.warning(self, "Ошибка", "Введите номер партии")
            return
        try:
            batch_index = int(number_str)
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Номер партии должен быть целым числом")
            return
        folder = self.get_batch_folder_by_index(batch_index)
        if not folder:
            QMessageBox.warning(self, "Ошибка", f"Партия с номером {batch_index} не найдена")
            return
        self.parse_final_report_and_plot(folder)

    def load_last_completed_batch(self):
        """Загружает последнюю завершённую партию."""
        folder = self.get_last_completed_batch_folder()
        if not folder:
            QMessageBox.warning(self, "Ошибка", "Нет завершённых партий с отчётами")
            return
        self.parse_final_report_and_plot(folder)

    def open_current_report(self):
        if not hasattr(self, 'current_report_path') or not self.current_report_path:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите партию")
            return
        report_path = Path(self.current_report_path)
        if not report_path.exists():
            QMessageBox.warning(self, "Ошибка", f"Файл отчёта не найден:\n{report_path}")
            return
        import os
        import sys
        if sys.platform == 'win32':
            os.startfile(report_path)
        elif sys.platform == 'darwin':
            import subprocess
            subprocess.call(['open', str(report_path)])
        else:
            import subprocess
            subprocess.call(['xdg-open', str(report_path)])