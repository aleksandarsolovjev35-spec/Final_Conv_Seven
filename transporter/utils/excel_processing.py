import csv
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

CLASS_NAMES_MAP = {
    "bottom_glass": "Стекло на дне",
    "glass": "Стекло",
    "welding": "Брак сварки",
    "uneven_heights": "Разновысотность",
    "window_sinks": "Раковина в окне",
    "flatness": "Плоскостность",
    "sinks": "Провал",
    "omission": "Пропуск",
    "Omission": "Пропуск",
    "mechanics": "Механика",
    "contacts": "Контакт",
    "contacts_long": "Длинный контакт",
    "flatness_short": "Короткая плоскостность",
    "platform": "Платформа",
    "objects": "Объекты",
    "2": "Разновысотность",
    "3": "Раковина/Провал",
    "4": "Стекло",
    "8": "Брак сварки",
    "1": "не брак",
    "0": "не брак",

}


def parse_timestamp(timestamp_str):
    """Преобразует timestamp из формата '2026-02-10_14-35-10-727' в 'ЧЧ:ММ:СС'"""
    try:
        parts = timestamp_str.split('_')
        if len(parts) == 2:
            time_part = parts[1]
            time_parts = time_part.split('-')
            if len(time_parts) >= 3:
                return f"{time_parts[0]}:{time_parts[1]}:{time_parts[2]}"
        return "??:??:??"
    except:
        return "??:??:??"


def parse_date(timestamp_str):
    """Преобразует '2026-02-10_14-35-10-727' в 'DD.MM.YYYY HH:MM'"""
    try:
        parts = timestamp_str.split('_')
        if len(parts) == 2:
            date_part = parts[0]
            time_part = parts[1]
            y, m, d = date_part.split('-')
            time_parts = time_part.split('-')
            if len(time_parts) >= 3:
                return f"{d}.{m}.{y} {time_parts[0]}:{time_parts[1]}"
        return "??.??.???? ??:??"
    except:
        return "??.??.???? ??:??"


def parse_defect_types_from_csv(defect_str):
    """Преобразует строку типа '[4]' или '[8]' в читаемый список."""
    if not defect_str or defect_str == "[]":
        return [0]
    try:
        # Убираем квадратные скобки и разбиваем по запятой
        inner = defect_str.strip("[]")
        if not inner:
            return []
        codes = [int(x.strip()) for x in inner.split(',')]
        return codes
    except:
        return []


def get_defect_codes_from_metadata(metadata_path):
    """Возвращает список кодов брака из метаданных (2,3,4,8)."""
    if not metadata_path.exists():
        return []
    try:
        with open(metadata_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except:
        return []
    codes = []
    cameras = data.get("cameras", {})
    for cam_data in cameras.values():
        objects = cam_data.get("objects", [])
        for obj in objects:
            cn = obj.get("class_name")
            if cn == "glass" or cn == "bottom_glass":
                codes.append(4)
            elif cn == "welding":
                codes.append(8)
            elif cn == "uneven_heights":
                codes.append(2)
            elif cn == "window_sinks":
                codes.append(3)
            # другие типы игнорируем
    return codes


def get_defect_types_from_metadata(metadata_path):
    """Извлекает уникальные class_name из метаданных и возвращает строку через запятую."""
    if not metadata_path.exists():
        return "файл не найден"
    try:
        with open(metadata_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except:
        return "ошибка чтения"
    class_names = set()
    cameras = data.get("cameras", {})
    for cam_data in cameras.values():
        objects = cam_data.get("objects", [])
        for obj in objects:
            cn = obj.get("class_name")
            if cn:
                class_names.add(cn)
    if not class_names:
        return ""
    readable = [CLASS_NAMES_MAP.get(cn, cn) for cn in sorted(class_names)]
    return ", ".join(readable)


def generate_report(batch_folder_path, output_path=None, max_items=None):
    batch_path = Path(batch_folder_path)
    if not batch_path.exists():
        raise FileNotFoundError(f"Папка партии не найдена: {batch_path}")

    stats_file = batch_path / "batch_statistics.json"
    if not stats_file.exists():
        raise FileNotFoundError(f"Файл статистики не найден: {stats_file}")
    with open(stats_file, 'r', encoding='utf-8') as f:
        stats = json.load(f)
    start_time_raw = stats.get("start_time", "")
    start_time_formatted = parse_date(start_time_raw) if start_time_raw else "??.??.???? ??:??"

    csv_file = batch_path / "screenshots_log.csv"
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV-лог не найден: {csv_file}")

    metadata_dir = batch_path / "metadata"

    items = []
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        all_rows = list(reader)
        if max_items is not None and max_items > 0:
            all_rows = all_rows[-max_items:]

        for i, row in enumerate(all_rows, start=1):
            if not row or len(row) < 4:
                continue
            timestamp = row[0]
            # Получаем коды из метаданных
            metadata_file = metadata_dir / f"{timestamp}.json"
            codes_meta = get_defect_codes_from_metadata(metadata_file)
            # Получаем коды из CSV (колонка defect_types, индекс 2)
            codes_csv = parse_defect_types_from_csv(row[2] if len(row) > 2 else "")
            # Объединяем (метаданные важнее, но если их нет – берём из CSV)
            codes = codes_meta if codes_meta else codes_csv

            # Определяем вердикт
            has_clean = any(c in (4, 8) for c in codes)
            has_defect = any(c in (2, 3) for c in codes)
            if not codes:
                verdict_str = "годен"
            elif has_defect:
                verdict_str = "брак"
            elif has_clean:
                verdict_str = "чистка"
            else:
                verdict_str = "годен"

            time_str = parse_timestamp(timestamp)

            # Получаем строку типов брака: сначала из метаданных, если пусто – из CSV
            defect_types_str = get_defect_types_from_metadata(metadata_file)
            if not defect_types_str and codes_csv:
                # Преобразуем коды из CSV в читаемые названия
                readable = []
                for code in sorted(set(codes_csv)):
                    readable.append(CLASS_NAMES_MAP.get(str(code), str(code)))
                defect_types_str = ", ".join(readable)

            items.append({
                "number": i,
                "result": verdict_str,
                "defect_types": defect_types_str,
                "time": time_str,
                "timestamp": timestamp
            })

    if not items:
        print("Нет данных в CSV")
        return None

    # Конечное время
    if all_rows:
        last_timestamp_raw = all_rows[-1][0]
        end_time_formatted = parse_date(last_timestamp_raw)
    else:
        end_time_formatted = "??.??.???? ??:??"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Стили
    title_font = Font(name='Calibri', size=14, bold=True)
    normal_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', size=11, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Заголовок
    ws.merge_cells('A1:G1')
    ws['A1'] = "Протокол инспекции № 000001"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    info_rows = [
        ("Номер рабочего места:", "?"),
        ("Тип изделия:", "?"),
        ("Идентификатор партии:", "?")
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = bold_font
        ws[f'B{i}'].font = normal_font

    ws['A7'] = "Начало инспекции:"
    ws['B7'] = start_time_formatted
    ws['A8'] = "Конец инспекции:"
    ws['B8'] = end_time_formatted
    ws['A7'].font = bold_font
    ws['A8'].font = bold_font

    ws['A10'] = ""

    # Заголовки таблицы
    headers = ["№ ячейки", "Результат проверки", "Тип брака", "Время измерения",
               "Фото Cam0", "Фото Cam1", "Фото Cam2"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Заполнение данных
    for i, item in enumerate(items, start=12):
        ws[f'A{i}'] = item["number"]
        ws[f'B{i}'] = item["result"]
        ws[f'C{i}'] = item["defect_types"]
        ws[f'D{i}'] = item["time"]

        # Цвет для результата
        if item["result"] == "брак":
            ws[f'B{i}'].font = Font(color="FF0000")
        elif item["result"] == "годен":
            ws[f'B{i}'].font = Font(color="008000")
        # Для "чистка" оставляем чёрный

        # Выравнивание времени
        ws[f'D{i}'].alignment = Alignment(horizontal='center')

        # Гиперссылки на фото
        for cam in range(3):
            cell = ws.cell(row=i, column=5 + cam)
            photo_path = f"Camera_{cam}/{item['timestamp']}.jpg"
            cell.hyperlink = photo_path
            cell.value = f"Фото {cam+1}"
            cell.font = Font(underline='single', color="0563C1")

    # Автоширина
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15
    ws.column_dimensions['C'].width = 35

    if output_path is None:
        output_path = batch_path / "report.xlsx"
    wb.save(output_path)
    return str(output_path)